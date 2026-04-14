"""
Converte um template Excel de Search Sets do Navisworks em XML.

Estrutura esperada do Excel:
- Aba "CONFIG": chave/valor a partir da linha 3
- Aba "SELECTION_SETS": dados a partir da linha 3

O XML gerado segue o mesmo padrao de "Search Set.xml".
"""

from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from uuid import uuid4
from xml.dom import minidom
from xml.etree import ElementTree as ET

import openpyxl


CONFIG_SHEET_NAME = "CONFIG"
SETS_SHEET_NAME = "SELECTION_SETS"
CONFIG_START_ROW = 3
SETS_START_ROW = 3


def _normalize_text(value: object) -> str:
    """Converte valores do Excel em texto sem ruido visual."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_config(path: str | Path) -> dict[str, str]:
    """Le a aba CONFIG e retorna um dicionario chave/valor."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[CONFIG_SHEET_NAME]
    config: dict[str, str] = {}

    for key, value in ws.iter_rows(min_row=CONFIG_START_ROW, values_only=True):
        key_text = _normalize_text(key)
        if not key_text:
            continue
        config[key_text] = _normalize_text(value)

    return config


def read_selection_sets(path: str | Path) -> "OrderedDict[str, list[dict[str, str]]]":
    """Le a aba SELECTION_SETS e agrupa as condicoes por selectionset_name."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SETS_SHEET_NAME]
    grouped: "OrderedDict[str, list[dict[str, str]]]" = OrderedDict()

    for row in ws.iter_rows(min_row=SETS_START_ROW, values_only=True):
        selectionset_name = _normalize_text(row[0])
        if not selectionset_name:
            continue

        grouped.setdefault(selectionset_name, []).append(
            {
                "condition_test": _normalize_text(row[1]) or "equals",
                "category_display": _normalize_text(row[2]),
                "category_internal": _normalize_text(row[3]),
                "property_display": _normalize_text(row[4]),
                "property_internal": _normalize_text(row[5]),
                "value_type": _normalize_text(row[6]) or "wstring",
                "value": _normalize_text(row[7]),
            }
        )

    return grouped


def build_search_sets_xml(
    config: dict[str, str],
    selectionsets: "OrderedDict[str, list[dict[str, str]]]",
) -> str:
    """Monta o XML final no formato de Search Set do Navisworks."""
    exchange = ET.Element(
        "exchange",
        attrib={
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsi:noNamespaceSchemaLocation": (
                "http://download.autodesk.com/us/navisworks/schemas/"
                "nw-exchange-12.0.xsd"
            ),
            "units": config.get("units", "mm"),
            "filename": config.get("filename", ""),
            "filepath": config.get("filepath", ""),
        },
    )

    selectionsets_el = ET.SubElement(exchange, "selectionsets")
    mode = config.get("mode", "all")
    disjoint = config.get("disjoint", "1")

    for set_name, conditions in selectionsets.items():
        selectionset_el = ET.SubElement(
            selectionsets_el,
            "selectionset",
            attrib={"name": set_name, "guid": str(uuid4())},
        )
        findspec_el = ET.SubElement(
            selectionset_el,
            "findspec",
            attrib={"mode": mode, "disjoint": disjoint},
        )
        conditions_el = ET.SubElement(findspec_el, "conditions")

        for index, condition in enumerate(conditions):
            condition_el = ET.SubElement(
                conditions_el,
                "condition",
                attrib={
                    "test": condition["condition_test"],
                    "flags": "0" if index == 0 else "64",
                },
            )

            category_el = ET.SubElement(condition_el, "category")
            ET.SubElement(
                category_el,
                "name",
                attrib={"internal": condition["category_internal"]},
            ).text = condition["category_display"]

            property_el = ET.SubElement(condition_el, "property")
            ET.SubElement(
                property_el,
                "name",
                attrib={"internal": condition["property_internal"]},
            ).text = condition["property_display"]

            value_el = ET.SubElement(condition_el, "value")
            ET.SubElement(
                value_el,
                "data",
                attrib={"type": condition["value_type"]},
            ).text = condition["value"]

        ET.SubElement(findspec_el, "locator").text = "/"

    raw_xml = ET.tostring(exchange, encoding="unicode")
    pretty_xml = minidom.parseString(raw_xml).toprettyxml(
        indent="  ",
        encoding="UTF-8",
    )
    lines = pretty_xml.decode("utf-8").splitlines()
    return "\n".join(line for line in lines if line.strip()) + "\n"


def convert_excel_to_xml(
    input_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    """Executa a conversao completa do Excel para XML."""
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else input_path.with_suffix(".xml")

    config = read_config(input_path)
    selectionsets = read_selection_sets(input_path)
    xml_content = build_search_sets_xml(config, selectionsets)
    output_path.write_text(xml_content, encoding="utf-8")
    return output_path


if __name__ == "__main__":
    default_input = Path("nw_search_template_v2.xlsx")
    output = convert_excel_to_xml(default_input)
    print(f"XML gerado em: {output}")
